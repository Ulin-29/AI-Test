# ==============================================================================
# BAGIAN 1: IMPORT
# ==============================================================================
import os
import shutil
import tempfile
import fitz 
import cv2
import asyncio
import random
import uuid
import json
import time
import requests
from datetime import datetime, timedelta
from io import BytesIO
from typing import List, Dict, Any
from concurrent.futures import ThreadPoolExecutor

# --- Import untuk FastAPI & Web ---
from fastapi import FastAPI, Request, Form, Depends, HTTPException, status, UploadFile, File, Body
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
from werkzeug.utils import secure_filename

# --- Import untuk Database & Autentikasi ---
from sqlalchemy.orm import Session
from sqlalchemy import func
from passlib.context import CryptContext
from . import models
from .database import engine, SessionLocal
from .email_utils import send_notification_email, send_register_email, send_password_changed_email, send_email_otp

# --- Import dari Modul AI & Logika Verifikasi Anda ---
from .Verifikasi_Fuzzy_Fix import VERIFICATION_TEMPLATES
from .modules.hybrid_verifier import verify_document_hybrid
from .modules.page_classifier import classify_page_by_keywords
from .modules.summarizer import generate_summary
from .modules.signature_detector import check_signatures_in_pdf 
from .modules.dl_classifier import predict_page_class

# --- Import Pustaka Tambahan dari ai-fx ---
from user_agents import parse
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from dotenv import load_dotenv
from captcha.image import ImageCaptcha

# Pustaka Opsional untuk OCR diperiksa di sini
try:
    import easyocr
    EASYOCR_AVAILABLE = True
except ImportError:
    EASYOCR_AVAILABLE = False
try:
    import torch
    TORCH_AVAILABLE = True
except ImportError:
    TORCH_AVAILABLE = False

# Muat environment variables
env_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), ".env")
load_dotenv()

# ==============================================================================
# BAGIAN 2: SETUP APLIKASI, DIREKTORI, DATABASE, & TEMPLATE
# ==============================================================================

# Buat tabel di database (jika belum ada)
models.Base.metadata.create_all(bind=engine)

# Inisialisasi aplikasi FastAPI
app = FastAPI(title="Document Verification AI")

# Tambahkan middleware untuk manajemen sesi (login, dll)
app.add_middleware(SessionMiddleware, secret_key=os.getenv("SECRET_KEY", "rahasia12345"))

# Dapatkan path absolut ke direktori tempat main.py berada
BASE_APP_DIR = os.path.dirname(os.path.abspath(__file__))
ASSETS_DIR = os.path.join(BASE_APP_DIR, "assets")
TEMPLATES_DIR = os.path.join(BASE_APP_DIR, "templates")

# Setup direktori untuk file statis dan template dengan path yang sudah pasti benar
app.mount("/assets", StaticFiles(directory=ASSETS_DIR), name="assets")
templates = Jinja2Templates(directory=TEMPLATES_DIR)

# Setup direktori upload
UPLOAD_DIR_FOTO = os.path.join(ASSETS_DIR, "uploads")
UPLOAD_DIR_DOKUMEN = os.path.join(BASE_APP_DIR, "dokumen_tersimpan") # Disimpan di luar assets
TEMP_UPLOADS_DIR = os.path.join(BASE_APP_DIR, "temp_uploads")
os.makedirs(UPLOAD_DIR_FOTO, exist_ok=True)
os.makedirs(UPLOAD_DIR_DOKUMEN, exist_ok=True)
os.makedirs(TEMP_UPLOADS_DIR, exist_ok=True)


# Setup untuk hashing password
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# Buat admin default saat aplikasi pertama kali jalan
def create_default_admin():
    db = SessionLocal()
    try:
        if not db.query(models.User).filter(models.User.email == "admin@mail.com").first():
            admin_user = models.User(
                email="admin@mail.com",
                username="Admin",
                password=pwd_context.hash("admin123")
            )
            db.add(admin_user)
            db.commit()
            print("✅ Admin user default berhasil dibuat.")
    finally:
        db.close()

create_default_admin()

# Dependency untuk mendapatkan sesi database
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# Dependency untuk mendapatkan user yang sedang login dari sesi
def get_current_user(request: Request, db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    if not user_id:
        return None
    return db.query(models.User).filter(models.User.id == user_id).first()


# ==============================================================================
# BAGIAN 3: FUNGSI HELPER UNTUK LOGIKA AI (OCR & VERIFIKASI)
# ==============================================================================
_easy_reader = None
def get_easyocr_reader(langs: List[str] = ['id', 'en'], force_gpu: bool = None):
    if not EASYOCR_AVAILABLE: raise RuntimeError("EasyOCR tidak tersedia.")
    global _easy_reader
    if _easy_reader is not None: return _easy_reader
    
    use_gpu = (TORCH_AVAILABLE and torch.cuda.is_available()) if force_gpu is None else force_gpu
    _easy_reader = easyocr.Reader(langs, gpu=use_gpu, verbose=False)
    print(f"✅ EasyOCR initialized | GPU={'Aktif' if use_gpu else 'Tidak Aktif (CPU Mode)'}")
    return _easy_reader

def easyocr_extract_text(image_path: str, min_conf: float = 0.25) -> str:
    try:
        reader = get_easyocr_reader()
        img_bgr = cv2.imread(image_path, cv2.IMREAD_COLOR)
        if img_bgr is None: return ""
        results = reader.readtext(img_bgr)
        return " ".join([t.strip() for (_, t, conf) in results if t and conf >= min_conf])
    except Exception as e:
        print(f"Error saat menjalankan OCR pada {image_path}: {e}")
        return ""

def compare_with_template_smart(pages_data: List[Dict[str, Any]], doc_type: str, signature_data: Dict[str, str]) -> List[Dict[str, Any]]:
    template = VERIFICATION_TEMPLATES.get(doc_type, [])
    if not template:
        return [{"name": "Error", "status": "TIDAK OK", "keterangan": "Tipe dokumen tidak valid.", "kategori": "Error"}]
        
    results = []
    detected_classes = {p['class'] for p in pages_data}
    
    class_map = {
        "Berita Acara Uji Terima (BAUT)": "BAUT", "Laporan Hasil Pekerjaan Selesai 100%": "LAPORAN_100%",
        "Surat Permintaan Uji Terima dari Mitra": "SURAT_PERMINTAAN", "SK/Penunjukan Team Uji Terima": "SK_TEAM",
        "Nota Dinas Pelaksanaan Uji Terima": "NOTA_DINAS", "Daftar Hadir Uji Terima": "DAFTAR_HADIR_UT",
        "BA Lapangan": "BA_LAPANGAN", "As-Built Drawing / Red Line Drawing": "RLD",
        "Bill of Quantity (BoQ) UT": "BOQ_UT", "Laporan Uji Terima": "LAPORAN_UT", "OTDR Report": "OTDR_REPORT",
        "Form OPM": "FORM_OPM", "Foto Kegiatan Uji Terima": "FOTO_KEGIATAN",
        "Foto Material terpasang sesuai BOQ": "FOTO_MATERIAL", "Foto Pengukuran OPM": "FOTO_PENGUKURAN_OPM",
        "Foto Roll Meter / Fault Locator": "FOTO_ROLL_METER", "Tanda Tangan Pejabat Berwenang": "SIGNATURE",
        "Berita Acara Test Commissioning (BACT)": "BACT", "Daftar Hadir Test Commissioning": "DAFTAR_HADIR_CT",
        "BOQ CT": "BOQ_CT", "Berita Acara Barang Tiba (BBA)": "BERITA_ACARA_BARANG_TIBA",
        "Red Line Drawing (RLD)": "RLD", "Foto Dokumentasi Test Comm": "FOTO_KEGIATAN",
        "Foto Capture Survey Address": "FOTO_SURVEY_ADDRESS", "Tanda Tangan Para Pihak": "SIGNATURE",
    }

    for item in template:
        name, kategori = item.get("name", "ITEM"), item.get("kategori", "-")
        if "Tanda Tangan" in name: continue
        target_class = class_map.get(name)
        found = target_class in detected_classes if target_class else False
        if not found and "Foto" in name and "EVIDENCE_PHOTO_UMUM" in detected_classes: found = True
        status, ket = ("OK", "DITEMUKAN") if found else ("TIDAK OK", "TIDAK DITEMUKAN")
        results.append({"name": name, "status": status, "keterangan": ket, "kategori": kategori})

    signature_item_name = "Tanda Tangan Pejabat Berwenang" if doc_type == "VERIFIKASI_BAUT" else "Tanda Tangan Para Pihak"
    
    # --- LOGIKA TANDA TANGAN DIKEMBALIKAN KE SEMULA ---
    ttd_status, ttd_ket = ("OK", "DITEMUKAN") if signature_data.get("status") == "Ditemukan" else ("TIDAK OK", "TIDAK DITEMUKAN")
        
    results.append({"name": signature_item_name, "kategori": "Validasi Akhir", "status": ttd_status, "keterangan": ttd_ket})
    return results

def process_verification_stream(pdf_path: str, doc_type: str):
    if not os.path.exists(pdf_path):
        yield {"status": "error", "message": "File tidak ditemukan"}
        return
    
    temp_dir = tempfile.mkdtemp()
    try:
        yield {"status": "processing", "message": "🚀 Membuka PDF...", "progress": 0}
        doc = fitz.open(pdf_path)
        total_pages = len(doc)
        yield {"status": "processing", "message": f"📄 Total halaman: {total_pages}", "progress": 5}
        
        yield {"status": "processing", "message": "🖊️ Mencari tanda tangan...", "progress": 10}
        # --- UBAH NAMA FUNGSI YANG DIPANGGIL KEMBALI KE check_signatures_in_pdf ---
        signature_results = check_signatures_in_pdf(pdf_path)
        yield {"status": "processing", "message": "✅ Pencarian tanda tangan selesai.", "progress": 15}

        temp_paths = [os.path.join(temp_dir, f"page_{i}.jpg") for i in range(total_pages)]
        for i, path in enumerate(temp_paths):
            doc.load_page(i).get_pixmap(dpi=200).save(path)
        doc.close()
        yield {"status": "processing", "message": "✅ PDF selesai dikonversi.", "progress": 20}

        def _classify_hybrid(path, index):
            p_class_dl, confidence = predict_page_class(path)
            text, p_class, p_class_kw = "", p_class_dl, "-"
            if confidence <= 0.75:
                text = easyocr_extract_text(path)
                p_class_kw = classify_page_by_keywords(text)
                if p_class_kw != "UNKNOWN": p_class = p_class_kw
            return {"class": p_class, "ai_class": p_class_dl, "keyword_class": p_class_kw, "path": path, "page_num": index + 1, "text": text}

        pages_data = []
        for i, path in enumerate(temp_paths):
            page_info = _classify_hybrid(path, i)
            pages_data.append(page_info)
            progress = int((((i + 1) / total_pages) * 70) + 20)
            print(f"Analisis Halaman {page_info['page_num']}/{total_pages} -> Keputusan Final: {page_info['class']}")
            yield {"status": "processing", "message": f"🤖 Menganalisis halaman {i + 1}/{total_pages}...", "progress": progress}
        
        yield {"status": "processing", "message": "🔍 Menyusun hasil akhir...", "progress": 95}
        
        results = compare_with_template_smart(pages_data, doc_type, signature_results)
        
        all_texts = [p['text'] or easyocr_extract_text(p['path']) for p in pages_data]
        summary = generate_summary(all_texts, [p['class'] for p in pages_data])
        ok = sum(1 for r in results if r["status"] == "OK")
        score = round(100 * ok / (len(results) or 1), 2)
        
        final_data = {"results": results, "score": score, "level": "Good" if score >= 70 else "Perlu Diperiksa", "summary": summary}
        yield {"status": "done", "data": final_data}

    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)
        
# ==============================================================================
# BAGIAN 4: RUTE (ENDPOINT) UNTUK OTENTIKASI & PENGGUNA
# ==============================================================================

# --- Lokasi & Captcha ---
@app.post("/update-location")
async def update_location(request: Request, data: dict = Body(...)):
    lat, lon = data.get("latitude"), data.get("longitude")
    if lat and lon:
        request.session["latitude"], request.session["longitude"] = lat, lon
    return {"status": "ok"}

@app.get("/captcha.png")
def get_captcha(request: Request):
    captcha_text = ''.join(random.choices('ABCDEFGHJKLMNPQRSTUVWXYZ23456789', k=5))
    request.session["captcha"] = captcha_text
    data = ImageCaptcha().generate(captcha_text)
    return StreamingResponse(data, media_type="image/png")

# --- Login & Logout ---
@app.get("/", response_class=HTMLResponse)
@app.get("/login", response_class=HTMLResponse)
def login_form(request: Request):
    message = request.session.pop("message", "")
    return templates.TemplateResponse("login.html", {"request": request, "message": message})

@app.post("/login", response_class=HTMLResponse)
async def login_post(request: Request, identifier: str = Form(...), password: str = Form(...), captcha_input: str = Form(...), db: Session = Depends(get_db)):
    identifier = identifier.strip().lower()
    email_error, password_error, captcha_error = "", "", ""

    if captcha_input.strip().upper() != request.session.get("captcha", ""):
        captcha_error = "Kode CAPTCHA salah."
    
    user = None
    if "@" in identifier:
        user = db.query(models.User).filter(func.lower(models.User.email) == identifier).first()
        if not user: email_error = "Email tidak ditemukan."
    elif identifier.isdigit():
        user = db.query(models.User).filter(models.User.phone == identifier).first()
        if not user: email_error = "Nomor HP tidak ditemukan."
    else:
        email_error = "Format email atau nomor HP tidak valid."

    if user and not pwd_context.verify(password, user.password):
        password_error = "Password salah."

    if email_error or password_error or captcha_error:
        return templates.TemplateResponse("login.html", {
            "request": request, "email": identifier, "email_error": email_error,
            "password_error": password_error, "captcha_error": captcha_error
        })

    # Login Berhasil
    request.session["user_id"], request.session["user_email"], request.session["username"], request.session["photo"] = user.id, user.email, user.username, user.photo
    
    try:
        ua_string = request.headers.get("user-agent", "")
        user_agent = parse(ua_string)
        device_name = f"{user_agent.os.family} - {user_agent.browser.family}"
        ip_address = request.client.host
        lat, lon = request.session.get("latitude"), request.session.get("longitude")
        lokasi, koordinat = "Lokasi tidak diketahui", ""

        if lat and lon: 
            koordinat = f"{lat},{lon}"
            try:
                url = f"https://nominatim.openstreetmap.org/reverse?format=json&lat={lat}&lon={lon}"
                res = requests.get(url, headers={"User-Agent": "LoginApp/1.0"}).json()
                lokasi = res.get("display_name", "Lokasi tidak diketahui")
            except Exception:
                lokasi = "Lokasi tidak diketahui (gagal geocode)"
        elif ip_address and not ip_address.startswith(("127.", "192.168.")):
            try:
                res = requests.get(f"https://ipinfo.io/{ip_address}/json").json()
                city, region = res.get("city", ""), res.get("region", "")
                lokasi = f"{city}, {region}" if city and region else "Lokasi dari IP"
                koordinat = res.get("loc", "")
            except Exception:
                lokasi = "Lokasi tidak diketahui (gagal IP lookup)"
        else:
            lokasi = "Jaringan Lokal"

        waktu_login = datetime.now().strftime("%A, %d %B %Y, %H:%M:%S")
        send_notification_email(to_email=user.email, subject="Notifikasi Login Berhasil", title=f"Halo {user.username},", message="Akun Anda baru saja berhasil login dari perangkat berikut:", device_name=device_name, waktu_login=waktu_login, lokasi=lokasi, ip_address=ip_address, koordinat=koordinat)
    except Exception as e:
        print(f"❌ GAGAL MENGIRIM NOTIFIKASI EMAIL: {e}")
        
    return RedirectResponse("/home", status_code=303)

@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/", status_code=302)

# --- Register ---
@app.get("/register", response_class=HTMLResponse)
def register_form(request: Request):
    return templates.TemplateResponse("register.html", {"request": request})

@app.post("/register", response_class=HTMLResponse)
def register_user(request: Request, db: Session = Depends(get_db), email: str = Form(...), phone_number: str = Form(...), username: str = Form(...), password: str = Form(...), confirm_password: str = Form(...)):
    email, phone_number = email.strip().lower(), phone_number.strip()
    
    if not email.endswith("@gmail.com"):
        return templates.TemplateResponse("register.html", {"request": request, "message": "Email harus menggunakan domain @gmail.com"})
    if len(password) < 6:
        return templates.TemplateResponse("register.html", {"request": request, "message": "Password harus minimal 6 karakter"})
    if password != confirm_password:
        return templates.TemplateResponse("register.html", {"request": request, "message": "Konfirmasi password tidak cocok"})
    if not username[0].isupper():
        return templates.TemplateResponse("register.html", {"request": request, "message": "Huruf pertama username harus kapital"})
    
    if db.query(models.User).filter((models.User.email == email) | (models.User.phone == phone_number) | (models.User.username == username)).first():
        return templates.TemplateResponse("register.html", {"request": request, "message": "Email, username, atau nomor HP sudah terdaftar"})
    
    new_user = models.User(email=email, phone=phone_number, username=username, password=pwd_context.hash(password))
    db.add(new_user)
    db.commit()
    send_register_email(email, username)
    request.session["message"] = "Akun berhasil dibuat. Silakan login."
    return RedirectResponse("/", status_code=302)

# --- Lupa Password & OTP ---
@app.get("/forgot-password", response_class=HTMLResponse)
async def forgot_password_form(request: Request):
    return templates.TemplateResponse("forgotpw.html", {"request": request})

@app.post("/forgot-password")
async def process_forgot_password_email(request: Request, email: str = Form(...), db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.email == email.strip().lower()).first()
    if not user:
        return templates.TemplateResponse("forgotpw.html", {"request": request, "error": "Email tidak ditemukan"})
    
    otp_code, expiry = str(random.randint(100000, 999999)), datetime.utcnow() + timedelta(minutes=5)
    user.otp_code, user.otp_expiry = otp_code, expiry
    db.commit()
    send_email_otp(user.email, otp_code)
    request.session["otp_user_id"] = user.id
    return RedirectResponse(url="/verify-otp", status_code=302)

@app.get("/verify-otp", response_class=HTMLResponse)
async def show_verify_otp(request: Request):
    return templates.TemplateResponse("verifyotp.html", {"request": request})

@app.post("/verify-otp")
async def process_verify_otp(request: Request, otp_code: str = Form(...), db: Session = Depends(get_db)):
    user_id = request.session.get("otp_user_id")
    if not user_id: return RedirectResponse(url="/forgot-password")
    user = db.query(models.User).get(user_id)

    if not user or otp_code.strip() != user.otp_code:
        return templates.TemplateResponse("verifyotp.html", {"request": request, "error": "Kode OTP salah."})
    if user.otp_expiry and datetime.utcnow() > user.otp_expiry:
        return templates.TemplateResponse("verifyotp.html", {"request": request, "error": "Kode OTP sudah kadaluarsa."})
    
    request.session["otp_verified"], request.session["user_id_for_reset"] = True, user.id
    user.otp_code, user.otp_expiry = None, None
    db.commit()
    return RedirectResponse(url="/reset-password", status_code=302)

# --- Kirim Ulang OTP  ---
@app.get("/resend-otp")
async def resend_otp(request: Request, db: Session = Depends(get_db)):
    user_id = request.session.get("otp_user_id")
    if not user_id: return RedirectResponse("/forgot-password")
    user = db.query(models.User).get(user_id)
    if not user: return RedirectResponse("/forgot-password")
    
    otp_code, expiry = str(random.randint(100000, 999999)), datetime.utcnow() + timedelta(minutes=5)
    user.otp_code, user.otp_expiry = otp_code, expiry
    db.commit()
    send_email_otp(user.email, otp_code)
    return RedirectResponse("/verify-otp", status_code=302)

# --- Ganti Password ---
@app.get("/reset-password", response_class=HTMLResponse)
async def show_reset_password(request: Request):
    if not (request.session.get("otp_verified") or request.session.get("password_verified")):
        return RedirectResponse("/", status_code=302)
    return templates.TemplateResponse("resetpw.html", {"request": request})

@app.post("/reset-password")
async def process_reset_password(request: Request, new_password: str = Form(...), confirm_password: str = Form(...), db: Session = Depends(get_db)):
    if not (request.session.get("otp_verified") or request.session.get("password_verified")):
        return RedirectResponse(url="/", status_code=302)
    
    if len(new_password) < 6:
        return templates.TemplateResponse("resetpw.html", {"request": request, "message": "Password harus minimal 6 karakter."})
    if new_password != confirm_password:
        return templates.TemplateResponse("resetpw.html", {"request": request, "message": "Password konfirmasi tidak cocok."})
    
    user_id = request.session.get("user_id_for_reset") or request.session.get("user_id")
    user = db.query(models.User).get(user_id)
    if not user: return RedirectResponse(url="/", status_code=302)
    
    user.password = pwd_context.hash(new_password)
    db.commit()
    send_password_changed_email(user.email, user.username)
    
    request.session.pop("otp_verified", None)
    request.session.pop("password_verified", None)
    request.session.pop("user_id_for_reset", None)

    return templates.TemplateResponse("success_redirect.html", {"request": request})

# ==============================================================================
# BAGIAN 5: RUTE HALAMAN UTAMA (SETELAH LOGIN)
# ==============================================================================

@app.get("/home", response_class=HTMLResponse)
def home(request: Request, db: Session = Depends(get_db), user: models.User = Depends(get_current_user)):
    if not user: 
        return RedirectResponse(url="/", status_code=302)

    # Mengambil 5 riwayat terbaru (
    riwayat_terbaru = db.query(models.Dokumen).filter(
        models.Dokumen.user_id == user.id
    ).order_by(models.Dokumen.timestamp.desc()).limit(5).all()

    # Hitung jumlah dokumen dengan status DITERIMA dan DITOLAK
    jumlah_berhasil = db.query(models.Dokumen).filter(
        models.Dokumen.user_id == user.id,
        models.Dokumen.status == "DITERIMA"
    ).count()

    jumlah_gagal = db.query(models.Dokumen).filter(
        models.Dokumen.user_id == user.id,
        models.Dokumen.status == "DITOLAK"
    ).count()

    return templates.TemplateResponse(
        "home.html", 
        {
            "request": request, 
            "user": user, 
            "now": datetime.now(),
            "riwayat_terbaru": riwayat_terbaru,
            "jumlah_berhasil": jumlah_berhasil,  # <-- Kirim data hitungan
            "jumlah_gagal": jumlah_gagal         # <-- Kirim data hitungan
        }
    )

@app.get("/riwayat", response_class=HTMLResponse)
def riwayat_page(request: Request, db: Session = Depends(get_db), user: models.User = Depends(get_current_user)):
    if not user: return RedirectResponse(url="/", status_code=302)
    
    # Ambil data riwayat dari database
    riwayat_dokumen = db.query(models.Dokumen).filter(models.Dokumen.user_id == user.id).order_by(models.Dokumen.timestamp.desc()).all()
    
    return templates.TemplateResponse("riwayat.html", {"request": request, "user": user, "riwayat_list": riwayat_dokumen})

@app.get("/api/riwayat/{dokumen_id}")
async def get_detail_riwayat_api(
    dokumen_id: int,
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user)
):
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")

    dokumen = db.query(models.Dokumen).filter(
        models.Dokumen.id == dokumen_id,
        models.Dokumen.user_id == user.id
    ).first()

    if not dokumen:
        raise HTTPException(status_code=404, detail="Dokumen tidak ditemukan")
    
    tipe_mapping = {
        "VERIFIKASI_BAUT": "Verifikasi Berita Acara Uji Terima (BAUT)",
        "VERIFIKASI_BACT": "Verifikasi Berita Acara Test Commissioning (BACT)"
    }
    # Gunakan mapping, atau tampilkan teks asli jika tidak ada di mapping
    tipe_dokumen_jelas = tipe_mapping.get(dokumen.tipe_dokumen, dokumen.tipe_dokumen)

    # Ubah data model ke format dictionary agar bisa dikirim sebagai JSON
    dokumen_data = {
        "nama_dokumen": dokumen.nama_dokumen,
        "tipe_dokumen": tipe_dokumen_jelas,
        "skor": dokumen.skor,
        "status": dokumen.status,
        "hasil_verifikasi": dokumen.hasil_verifikasi,
        "ringkasan": dokumen.ringkasan
    }
    return JSONResponse(content=dokumen_data)

# ======================== PROFILE (SESUAI AI-FX ASLI) ========================
@app.get("/profile", response_class=HTMLResponse)
def profile_page(request: Request, user: models.User = Depends(get_current_user)):
    if not user: return RedirectResponse(url="/", status_code=302)
    message = request.session.pop("temp_message", None)
    return templates.TemplateResponse("profile.html", {"request": request, "user": user, "message": message})

@app.get("/edit-profile", response_class=HTMLResponse)
def edit_profile_page(request: Request, user: models.User = Depends(get_current_user)):
    if not user: return RedirectResponse(url="/", status_code=302)
    message = request.session.pop("temp_message", None)
    return templates.TemplateResponse("edit-profile.html", {"request": request, "user": user, "message": message})

@app.post("/update-profile", response_class=HTMLResponse)
async def update_profile(request: Request, db: Session = Depends(get_db), user: models.User = Depends(get_current_user), username: str = Form(...), email: str = Form(...), photo: UploadFile = File(None)):
    if not user: return RedirectResponse("/login")
    
    email = email.strip().lower()
    if not username[0].isupper():
        return templates.TemplateResponse("edit-profile.html", {"request": request, "user": user, "message": "Huruf pertama username harus kapital"})
    if db.query(models.User).filter(models.User.username == username, models.User.id != user.id).first():
        return templates.TemplateResponse("edit-profile.html", {"request": request, "user": user, "message": "Username sudah digunakan."})
    if db.query(models.User).filter(models.User.email == email, models.User.id != user.id).first():
        return templates.TemplateResponse("edit-profile.html", {"request": request, "user": user, "message": "Email sudah digunakan."})

    user.username, user.email = username, email

    if photo and photo.filename:
        if user.photo and user.photo != "default.png":
            old_path = os.path.join(UPLOAD_DIR_FOTO, user.photo)
            if os.path.exists(old_path): os.remove(old_path)
        new_filename = f"user_{user.id}_{int(time.time())}{os.path.splitext(photo.filename)[1]}"
        save_path = os.path.join(UPLOAD_DIR_FOTO, new_filename)
        with open(save_path, "wb") as f: f.write(await photo.read())
        user.photo = new_filename

    db.commit()
    db.refresh(user)
    request.session["user_email"], request.session["username"], request.session["photo"] = user.email, user.username, user.photo
    return templates.TemplateResponse("profile.html", {"request": request, "user": user, "message": "Profil berhasil diperbarui."})

@app.post("/delete-photo")
async def delete_profile_photo(request: Request, db: Session = Depends(get_db), user: models.User = Depends(get_current_user)):
    if not user: return RedirectResponse(url="/")
    if user.photo and user.photo != "default.png":
        file_path = os.path.join(UPLOAD_DIR_FOTO, user.photo)
        if os.path.exists(file_path): os.remove(file_path)
        user.photo, request.session["photo"] = "default.png", "default.png"
        db.commit()
        request.session["temp_message"] = "Foto profil berhasil dihapus."
    else:
        request.session["temp_message"] = "Tidak ada foto untuk dihapus."
    return RedirectResponse(url="/edit-profile", status_code=302)

# ======================== VERIFY PASSWORD ========================
@app.get("/verifypw", response_class=HTMLResponse)
async def show_verify_password(request: Request, user: models.User = Depends(get_current_user)):
    if not user: return RedirectResponse("/", status_code=302)
    return templates.TemplateResponse("verifypw.html", {"request": request})

@app.post("/verifypw", response_class=HTMLResponse)
async def verify_password(request: Request, current_password: str = Form(...), db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    if not user_id: return RedirectResponse("/", status_code=302)
    user = db.query(models.User).get(user_id)
    if not user or not pwd_context.verify(current_password, user.password):
        return templates.TemplateResponse("verifypw.html", {"request": request, "error_message": "Password salah. Coba lagi."})
    
    request.session["password_verified"] = True
    request.session["user_id_for_reset"] = user.id
    return RedirectResponse("/reset-password", status_code=302)

# ==============================================================================
# BAGIAN 6: RUTE (ENDPOINT) UNTUK PROSES VERIFIKASI AI
# ==============================================================================
@app.get("/verifikasi", response_class=HTMLResponse)
def verifikasi_page(request: Request, user: models.User = Depends(get_current_user)):
    if not user: return RedirectResponse(url="/", status_code=302)
    return templates.TemplateResponse("verifikasi.html", {"request": request, "user": user})

@app.post("/verify-stream")
async def verify_document_stream_endpoint(
    request: Request,
    doc_type: str = Form(...),
    file: UploadFile = File(...),
    unique_filename: str = Form(...),
    user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not user: raise HTTPException(status_code=401, detail="Not authenticated")

    save_path = os.path.join(UPLOAD_DIR_DOKUMEN, unique_filename)
    with open(save_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    async def event_generator():
        final_result_data = None
        try:
            gen = process_verification_stream(save_path, doc_type)
            for update in gen:
                if await request.is_disconnected():
                    # Jika koneksi terputus, picu CancelledError
                    raise asyncio.CancelledError()

                if update.get("status") == "done":
                    final_result_data = update.get("data")

                yield f"data: {json.dumps(update)}\n\n"
                await asyncio.sleep(0.02)

            # >> BLOK PENYIMPANAN KE DATABASE <<
            if final_result_data:
                score = final_result_data.get("score", 0)
                status_dokumen = "DITERIMA" if score >= 70 else "DITOLAK"
                dokumen_baru = models.Dokumen(
                    nama_dokumen=file.filename,
                    nama_file_unik=unique_filename,
                    tipe_dokumen=doc_type,
                    status=status_dokumen,
                    skor=int(score),
                    hasil_verifikasi=final_result_data.get("results"),
                    ringkasan=final_result_data.get("summary"),
                    user_id=user.id
                )
                db.add(dokumen_baru)
                db.commit()
                print(f"✅ Hasil verifikasi untuk {unique_filename} berhasil disimpan ke DB.")

        except asyncio.CancelledError:
            # >> BLOK PEMBERSIHAN FILE <<
            # Berjalan HANYA jika proses dibatalkan.
            print(f"⚠️ Proses dibatalkan untuk {unique_filename}. File akan dihapus.")
            if os.path.exists(save_path):
                os.remove(save_path)
                print(f"🗑️ File terhapus karena proses dibatalkan: {save_path}")
        
        except Exception as e:
            print(f"Error dalam event_generator: {e}")
            if os.path.exists(save_path):
                os.remove(save_path)
                print(f"🗑️ File terhapus karena error: {save_path}")

    return StreamingResponse(event_generator(), media_type="text/event-stream")

@app.post("/cancel-verification")
async def cancel_verification(request: Request, data: dict = Body(...)):
    """Menerima permintaan dari klien untuk menghapus file yang prosesnya dibatalkan."""
    filename = data.get("filename")
    if not filename:
        raise HTTPException(status_code=400, detail="Nama file tidak disertakan.")

    file_path = os.path.join(UPLOAD_DIR_DOKUMEN, os.path.basename(filename))

    try:
        if os.path.exists(file_path):
            os.remove(file_path)
            print(f"🗑️ File yang dibatalkan telah dihapus: {file_path}")
            return JSONResponse(content={"status": "success", "message": "File berhasil dihapus."}, status_code=200)
        else:
            # Ini bisa terjadi jika pembatalan sangat cepat.
            print(f"⚠️ File yang akan dibatalkan tidak ditemukan (mungkin sudah dihapus): {file_path}")
            return JSONResponse(content={"status": "not_found", "message": "File tidak ditemukan."}, status_code=404)
    except Exception as e:
        print(f"❌ Gagal menghapus file yang dibatalkan: {e}")
        raise HTTPException(status_code=500, detail=f"Gagal menghapus file: {e}")

@app.post("/hapus-dokumen/{dokumen_id}")
async def hapus_dokumen(request: Request, dokumen_id: int, db: Session = Depends(get_db), user: models.User = Depends(get_current_user)):
    if not user:
        return RedirectResponse(url="/", status_code=302)

    # 1. Cari dokumen di database berdasarkan ID uniknya
    dokumen_to_delete = db.query(models.Dokumen).filter(
        models.Dokumen.id == dokumen_id, 
        models.Dokumen.user_id == user.id # Pastikan user hanya bisa hapus miliknya
    ).first()

    if not dokumen_to_delete:
        # Jika dokumen tidak ditemukan, kembali saja
        return RedirectResponse(url="/riwayat", status_code=302)
    
    # 2. Ambil nama file unik dari database
    nama_file_di_server = dokumen_to_delete.nama_file_unik
    file_path = os.path.join(UPLOAD_DIR_DOKUMEN, nama_file_di_server)

    # 3. Hapus file fisik dari server
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
            print(f"🗑️ File fisik dihapus: {file_path}")
    except Exception as e:
        print(f"⚠️ Gagal menghapus file fisik {file_path}: {e}")
        # Anda bisa menambahkan pesan error untuk pengguna di sini jika perlu

    # 4. Hapus catatan dari database
    db.delete(dokumen_to_delete)
    db.commit()

    # 5. Kembalikan pengguna ke halaman riwayat
    return RedirectResponse(url="/riwayat", status_code=302)

@app.post("/download-excel")
async def download_excel_report(results: List[dict] = Body(...)):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Laporan Verifikasi"

    headers = ["NO", "ITEM YANG DIPERIKSA", "STATUS OK", "STATUS NOK", "KETERANGAN"]
    sheet.append(headers)

    # --- STYLING LEBAR KOLOM & TINGGI HEADER ---
    sheet.column_dimensions['A'].width = 5
    sheet.column_dimensions['B'].width = 50
    sheet.column_dimensions['C'].width = 12
    sheet.column_dimensions['D'].width = 12
    sheet.column_dimensions['E'].width = 40

    # BARU: Atur tinggi baris header agar lebih lega
    sheet.row_dimensions[1].height = 25

    # Atur header menjadi tebal dan rata tengah (vertikal & horizontal)
    bold_font = Font(bold=True)
    center_alignment_header = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for cell in sheet[1]:
        cell.font = bold_font
        cell.alignment = center_alignment_header
    
    # --- Mengisi data ke dalam baris ---
    for i, item in enumerate(results, 1):
        ok_tick = "✔" if item.get("status") == "OK" else ""
        nok_tick = "✔" if item.get("status") != "OK" else ""
        sheet.append([
            i, 
            item.get("name"), 
            ok_tick, 
            nok_tick, 
            item.get("keterangan")
        ])
    
    # --- STYLING BORDER & ALIGNMENT UNTUK SEMUA DATA ---
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    # BARU: Definisikan alignment untuk sel data
    center_alignment_data = Alignment(horizontal='center', vertical='center')
    left_alignment_data = Alignment(horizontal='left', vertical='center', wrap_text=True)

    max_row = sheet.max_row
    for row_index in range(2, max_row + 1): # Mulai dari baris 2 (data)
        sheet.row_dimensions[row_index].height = 25 # Atur tinggi standar untuk baris data

    for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border
            # Terapkan alignment berbeda untuk header dan data
            if cell.row > 1: # Jika bukan baris header
                if cell.column_letter in ['A', 'C', 'D']:
                    cell.alignment = center_alignment_data
                else:
                    cell.alignment = left_alignment_data
    
    # --- Proses untuk membuat dan mengirim file Excel ---
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": "attachment; filename=laporan_verifikasi.xlsx"}
    )